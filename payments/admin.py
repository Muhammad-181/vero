# admin.py

import openpyxl
from openpyxl.drawing.image import Image as XLImage
from django.http import HttpResponse
from django.contrib import admin
from .models import Payment, Fee, FeeAmount
from students.models import StudentProfile
from django.utils.timezone import localtime
from io import BytesIO
from PIL import Image as PILImage
from reportlab.lib.pagesizes import letter
from reportlab.pdfgen import canvas

class PaymentAdmin(admin.ModelAdmin):
    list_display = ('user', 'fee', 'reference', 'verified', 'timestamp')
    actions = ['export_as_excel']

    def export_as_excel(self, request, queryset):
        # Filter only verified payments
        queryset = queryset.filter(verified=True)

        # Create an Excel workbook and sheet
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Payments"

        # Define the column headers
        columns = ['User', 'Fee', 'Verified', 'Timestamp', 'Profile Picture', 'First Name', 'Last Name', 'Phone Number', 'Institution Email']
        ws.append(columns)

        for payment in queryset:
            student_profile = StudentProfile.objects.filter(user=payment.user).first()

            # Add payment and student profile details to the worksheet
            row = [
                payment.user.username,
                payment.fee.name,
                payment.verified,
                localtime(payment.timestamp).strftime('%Y-%m-%d %H:%M:%S'),
                '',  # Placeholder for the image
                student_profile.first_name if student_profile else '',
                student_profile.last_name if student_profile else '',
                student_profile.phone_number if student_profile else '',
                student_profile.inst_email if student_profile else ''
            ]
            ws.append(row)

            # Add the profile picture if it exists
            if student_profile and student_profile.profile_picture:
                image_path = student_profile.profile_picture.path
                img = XLImage(image_path)
                img.anchor = ws.cell(row=ws.max_row, column=6).coordinate  # Place the image in the 'Profile Picture' column
                ws.add_image(img)

        # Create an HTTP response with the Excel file
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename="payments_with_profiles.xlsx"'

        # Save the workbook to the response
        wb.save(response)
        return response
    

    export_as_excel.short_description = "Export Selected Payments to Excel"


admin.site.register(Payment, PaymentAdmin)

admin.site.register(Fee)
admin.site.register(FeeAmount)